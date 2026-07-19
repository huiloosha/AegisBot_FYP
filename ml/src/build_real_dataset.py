"""
build_real_dataset.py
---------------------
Convert the real user survey (Google Forms export, .xlsx) into a training CSV
in the EXACT format train.py expects: the 20 question codes (0-4), the 8
password features, and a rubric-computed `risk_level` label.

Per the supervisor's guidance:
  - TRAINING data  = these real survey responses (features real, labels by rubric)
  - TESTING data   = the synthetic dataset (used separately as a held-out test set)

So this script produces ONLY the real training set. Labels are derived with the
same evidence-based rubric used for the synthetic data, so the two are directly
comparable.

Usage:
    python build_real_dataset.py --in ../data/survey.xlsx --out ../data/AegisBot_Real_Training_Dataset.csv

What it does:
  1. Drops respondents who did not consent.
  2. Maps each free-text / numeric survey answer to the model's 0-4 scale.
     (Direction — direct/reverse — is NOT applied here; the rubric applies it
      when scoring, exactly as for the synthetic data.)
  3. Derives the 8 password features from the 3 password survey columns.
  4. Computes behaviour_score, password_score, overall risk, and risk_level
     using the rubric (reusing the same formulas as build_dataset.py).
  5. Writes a CSV with columns: assessment_id, <20 question codes>,
     <8 password features>, plus label columns (risk_level etc.).
"""
import argparse
import sys
from pathlib import Path

import numpy as np
import pandas as pd
import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))
from rubric import (QUESTIONS, CATEGORY_WEIGHTS, PASSWORD_RUBRIC,
                    OVERALL_WEIGHTS, risk_level_from_score)


# ---------------------------------------------------------------------------
# 1. Column mapping: survey column index -> question code
#    (based on the Google Form column order in the export)
# ---------------------------------------------------------------------------
# Survey columns 5..23 are 19 behavioural questions. NOTE: the survey does NOT
# include PM05 ("write down / share passwords"); it is imputed below as a safe
# default (0 = never), and this is recorded as a data limitation in the report.
COL_TO_CODE = {
    5:  "PM01",   6:  "PM02",   7:  "PM03",   8:  "PM04",
    9:  "AUTH01", 10: "AUTH02", 11: "AUTH03", 12: "AUTH04", 13: "AUTH05",
    14: "PHISH01", 15: "PHISH02", 16: "PHISH03", 17: "PHISH04", 18: "PHISH05",
    19: "SOC01",  20: "SOC02",  21: "SOC03",  22: "SOC04",  23: "SOC05",
}
# Questions not present in the survey -> imputed with a neutral/safe default (0).
IMPUTED_CODES = {"PM05": 0}
# Password survey columns:
COL_PW_LENGTH   = 24   # "Roughly how long is a typical password" -> band
COL_PW_CHARSET  = 25   # "Which of these does a typical password include" -> multiselect
COL_PW_COMMON   = 26   # "contains a common word/name/sequence?" -> Yes/No
COL_CONSENT     = 1
COL_TIMESTAMP   = 0


# ---------------------------------------------------------------------------
# 2. Answer-string -> 0-4 mapping
# ---------------------------------------------------------------------------
# Frequency words. Note the data contains a typo "ALways" -> handle it.
FREQUENCY_MAP = {
    "never": 0, "rarely": 1, "sometimes": 2, "often": 3, "always": 4,
    "usually": 3,                       # in case "Usually" appears
}
# 1-5 numeric scales (complexity, confidence, comfort) -> 0-4
# 1->0, 2->1, 3->2, 4->3, 5->4  (shift down by one)
def numeric_1to5_to_0to4(v):
    n = int(round(float(v)))
    return max(0, min(n - 1, 4))

# Yes/No -> 4/0 (position on the 0-4 axis; rubric applies direction)
YESNO_MAP = {"yes": 4, "no": 0}


def answer_to_04(raw, response_type):
    """Map one survey answer to 0-4.

    IMPORTANT: the real survey asked several questions the rubric labels as
    'yes_no' using frequency scales instead (e.g. PM03 'use a password manager'
    was answered Never/Sometimes/Always). So we map by the ANSWER CONTENT, not
    the rubric's declared response_type:
      - a real Yes/No answer   -> 4 / 0
      - a 1-5 numeric answer   -> 0-4 (shift down one)
      - a frequency word       -> Never..Always = 0..4
    The rubric applies direction (direct/reverse) later during scoring.
    """
    if raw is None:
        return None
    s = str(raw).strip().lower()

    # explicit yes/no
    if s in ("yes", "no"):
        return YESNO_MAP[s]

    # numeric 1-5 scale
    try:
        return numeric_1to5_to_0to4(float(s))
    except ValueError:
        pass

    # frequency word (handles the "ALways" typo via lowercasing)
    return FREQUENCY_MAP.get(s, None)


# ---------------------------------------------------------------------------
# 3. Password-feature derivation
# ---------------------------------------------------------------------------
def length_band_to_number(raw):
    """'Under 8 characters' -> 6, '8-11' -> 9, '12-15' -> 13, '16+' -> 17."""
    if raw is None:
        return 8
    s = str(raw).strip().lower()
    if "under 8" in s or s.startswith("<"):
        return 6
    if "8-11" in s or "8–11" in s:
        return 9
    if "12-15" in s or "12–15" in s:
        return 13
    if "16" in s:
        return 17
    # fallback: try to parse a leading number
    for tok in s.replace("-", " ").split():
        if tok.isdigit():
            return int(tok)
    return 8


def charset_flags(raw):
    """Parse the multiselect into has_upper/lower/number/symbol booleans."""
    s = str(raw or "").lower()
    return {
        "has_uppercase": "upper" in s,
        "has_lowercase": "lower" in s,
        "has_number":    "number" in s or "digit" in s,
        "has_symbol":    "symbol" in s,
    }


def common_word_flag(raw):
    return str(raw or "").strip().lower().startswith("y")


def derive_password_features(length_raw, charset_raw, common_raw):
    """Produce the 8 model password features from the 3 survey columns,
    using the SAME entropy formula as build_dataset.sample_password_features."""
    length = length_band_to_number(length_raw)
    cf = charset_flags(charset_raw)
    common = common_word_flag(common_raw)
    # Lowercase is assumed present if nothing was ticked (virtually all passwords).
    if not any(cf.values()):
        cf["has_lowercase"] = True

    # entropy ~ length * log2(charset), degraded by common pattern (same as synthetic)
    charset = 26 * (1 if cf["has_lowercase"] else 0)
    charset += 26 if cf["has_uppercase"] else 0
    charset += 10 if cf["has_number"] else 0
    charset += 32 if cf["has_symbol"] else 0
    charset = min(max(charset, 26), 94)
    entropy = length * np.log2(charset)
    if common:
        entropy *= 0.25
    entropy = round(float(entropy), 1)

    return {
        "password_length": int(length),
        "estimated_entropy": entropy,
        "has_uppercase": bool(cf["has_uppercase"]),
        "has_lowercase": bool(cf["has_lowercase"]),
        "has_number": bool(cf["has_number"]),
        "has_symbol": bool(cf["has_symbol"]),
        "common_pattern_detected": bool(common),
        # Survey doesn't ask about repeated characters; default False (conservative).
        "repeated_characters": False,
    }


# ---------------------------------------------------------------------------
# 4. Rubric scoring (same math as build_dataset.py, applied row-wise)
# ---------------------------------------------------------------------------
def score_row(answers, pw):
    """Return (behaviour_score, password_score, overall_risk, risk_level)."""
    # behaviour
    risk = 0.0
    for cat, cat_w in CATEGORY_WEIGHTS.items():
        cat_qs = [q for q in QUESTIONS if q["category"] == cat]
        s = 0.0
        for q in cat_qs:
            raw = float(answers[q["code"]])
            norm = raw / 4.0
            rf = norm if q["risk_direction"] == "direct" else (1 - norm)
            s += rf * q["weight"] * 100.0
        risk += s * cat_w
    behaviour_risk = max(0.0, min(risk, 100.0))
    behaviour_score = 100.0 - behaviour_risk

    # password (same as score_password)
    r = PASSWORD_RUBRIC
    strength = min(max(pw["estimated_entropy"] / r["entropy_scale_bits"] * 100.0, 0), 100)
    diversity = sum(int(pw[k]) for k in ("has_uppercase", "has_lowercase",
                                         "has_number", "has_symbol"))
    strength += min(diversity * r["diversity_bonus_each"], r["diversity_bonus_max"])
    if pw["common_pattern_detected"]:
        strength -= r["common_pattern_penalty"]
    if pw["repeated_characters"]:
        strength -= r["repeated_char_penalty"]
    if pw["password_length"] < r["short_length_threshold"]:
        strength -= r["short_length_penalty"]
    password_score = min(max(strength, 0), 100)
    password_risk = 100 - password_score

    # overall (same weighting as build_dataset)
    overall = (OVERALL_WEIGHTS["behaviour"] * behaviour_risk +
               OVERALL_WEIGHTS["password"] * password_risk)
    level = risk_level_from_score(overall)
    return round(behaviour_score, 1), round(password_score, 1), round(overall, 1), level


# ---------------------------------------------------------------------------
# 5. Main conversion
# ---------------------------------------------------------------------------
def convert(in_path, out_path):
    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    q_by_code = {q["code"]: q for q in QUESTIONS}
    records = []
    skipped_consent = 0
    skipped_incomplete = 0

    for r in rows:
        # consent gate
        consent = str(r[COL_CONSENT] or "").strip().lower()
        if not consent.startswith("y"):
            skipped_consent += 1
            continue

        # map the 20 answers
        answers = {}
        incomplete = False
        for col, code in COL_TO_CODE.items():
            rtype = q_by_code[code]["response_type"]
            val = answer_to_04(r[col], rtype)
            if val is None:
                incomplete = True
                break
            answers[code] = val
        if incomplete:
            skipped_incomplete += 1
            continue

        # add imputed answers for questions not present in the survey
        for code, default in IMPUTED_CODES.items():
            answers[code] = default

        # password features
        pw = derive_password_features(r[COL_PW_LENGTH], r[COL_PW_CHARSET], r[COL_PW_COMMON])

        # labels via rubric
        b_score, p_score, overall, level = score_row(answers, pw)

        rec = {"assessment_id": len(records) + 1}
        rec.update(answers)
        rec.update(pw)
        rec["behaviour_score"] = b_score
        rec["password_score"] = p_score
        rec["overall_risk_score"] = overall
        rec["risk_level"] = level
        rec["is_synthetic"] = False
        records.append(rec)

    df = pd.DataFrame(records)

    # Column order: assessment_id, questions, password features, labels
    q_cols = [q["code"] for q in QUESTIONS]
    pw_cols = ["password_length", "estimated_entropy", "has_uppercase",
               "has_lowercase", "has_number", "has_symbol",
               "common_pattern_detected", "repeated_characters"]
    label_cols = ["behaviour_score", "password_score", "overall_risk_score",
                  "risk_level", "is_synthetic"]
    df = df[["assessment_id"] + q_cols + pw_cols + label_cols]

    df.to_csv(out_path, index=False)

    # summary
    print(f"Wrote {len(df)} real training rows -> {out_path}")
    print(f"  skipped (no consent): {skipped_consent}")
    print(f"  skipped (incomplete): {skipped_incomplete}")
    print("  risk_level distribution:")
    print(df["risk_level"].value_counts().to_string())
    return df


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="inp", required=True, help="survey .xlsx path")
    ap.add_argument("--out", dest="out",
                    default="../data/AegisBot_Real_Training_Dataset.csv")
    args = ap.parse_args()
    convert(args.inp, args.out)
